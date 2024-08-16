import os
import pandas as pd
import numpy as np
import yfinance as yf
import xlsxwriter as xlsx
import openpyxl
from datetime import date
from modify_data.get_start_date import get_start_date
from modify_data.init_stock_data import init_stock_data
from modify_data.update_stock_data import update_stock_data


def init_excel_file():
    if not os.path.exists('stock_data.xlsx'):
        workbook = xlsx.Workbook('stock_data.xlsx')
        workbook.close()
#init_stock_data(stock_name, start_date, end_date)

# check for sheet name with stock_name
# if found, get start_date from that sheet
# if not found, init_date with fixed start_date of 2020-01-01
end_date = date.today()
stock_names = ['^GSPC', 'AAPL']
file_path = 'stock_data.xlsx'
init_excel_file()
workbook = openpyxl.load_workbook(file_path)
for stock_name in stock_names:
    if stock_name in workbook.sheetnames:
        print(f"'{stock_name}' 시트가 존재합니다.")
        start_date = get_start_date(stock_name)
        update_stock_data(file_path, stock_name, start_date, end_date)
    else:
        print(f"'{stock_name}' 시트가 존재하지 않습니다.")
        init_stock_data(stock_name, '2020-01-01', '2024-07-01')
